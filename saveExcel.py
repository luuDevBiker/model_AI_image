import cv2
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side
from pathlib import Path
import shutil
import os


def create_excel_file(arr_data):
    print('Creating workbook...')
    wb = Workbook()
    ws = wb.active
    ws.title = "Điểm thi"
    header_arr = ["Số Thứ tự", "Mã Sinh viên", "Họ và Tên", "Lớp", "Ký tên", "Ảnh Document", "Document",
                  "Ảnh Presentation", "Presentation"]
    # row_code = [code for code in range(ord('a'), ord('i') + 1)]
    # Write header to wb
    print('Writing header')
    for i in range(0, len(header_arr)):
        # print(header_arr[i])
        # print(get_column_letter(i))
        ws[get_column_letter(i + 1) + '1'] = header_arr[i]
    # get width of header row
    column_widths = []
    for i, cell in enumerate(header_arr):
        # print('i: ' + str(i))
        if len(column_widths) > i:
            if len(cell) > column_widths[i]:
                column_widths[i] = len(cell)
        else:
            column_widths += [len(cell)]
    print(column_widths)
    print('Writing data')
    # Write data to wb
    try:
        dirpath = Path('cell_img')
        if dirpath.exists() and dirpath.is_dir():
            temp_path = Path('cell_img/')
            shutil.rmtree(temp_path)
        dirpath.mkdir()
        # max_cell_width = max_cell_height = 0
        for i in range(len(arr_data)):
            # print('i: ' + str(i))
            # print('------------')
            for j in range(0, len(header_arr)):
                # print('j: ' + str(j))
                cell_name = get_column_letter(j + 1) + str(i + 2)
                # print(cell_name)
                if j != 6 and j != 8:
                    # print('anh')
                    # print(arr_data[i-2][j-1])
                    cell_img = arr_data[i][j]
                    # cv2.imshow(cell_name, cell_img)
                    # cv2.waitKey(0)
                    # print(cell_img.shape)
                    temp_height, temp_width, _ = cell_img.shape
                    img_name = str(i + 2) + '-' + str(j) + '.jpg'
                    img_path = 'cell_img/' + img_name
                    print('Img write : ' + str(cv2.imwrite(img_path, cell_img)))  # Ghi ảnh
                    cell_img = img_path

                    print('col w: ' + str(column_widths[j]) + ' img w: ' + str(temp_width))
                    if column_widths[j] < temp_width:
                        column_widths[j] = temp_width
                    print('new col w: ' + str(column_widths[j]))
                    # if max_cell_width < temp_width:
                    #     max_cell_width = temp_width
                    # if max_cell_height < temp_height:
                    #     max_cell_height = temp_height

                    # cv2.imshow(cell_name, cell_img)
                    # cv2.waitKey(0)

                    cell = Image(cell_img)
                    # print('created image')
                    ws.add_image(cell, cell_name)
                    # print('add to ws')
                else:
                    # print('so')
                    cell_data = arr_data[i][j]
                    ws[cell_name] = cell_data
                    # print(cell_name + ': ' + cell_data)

        # print('max_cell_width: ' + str(max_cell_width) + ' max_cell_height: ' + str(max_cell_height))

        print('Resize row and column...')
        print('col wd: ')
        print(column_widths)
        print('---------')
        for i in range(len(header_arr)):  # ,1 to start at 1
            print('column: ' + get_column_letter(i + 1))
            print(column_widths[i])
            if i != 6 and i != 8:
                ws.column_dimensions[get_column_letter(i + 1)].width = column_widths[i] / 7
            else:
                ws.column_dimensions[get_column_letter(i + 1)].width = column_widths[i] + 10
        for i in range(len(arr_data)):
            # print('row: ' + str(i))
            ws.row_dimensions[i + 2].height = 160 * 3 / 4 + 10
        for row in ws.iter_rows():
            for cell in row:
                # print('alig center')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(border_style='thin', color='000000'),
                                     right=Side(border_style='thin', color='000000'),
                                     top=Side(border_style='thin', color='000000'),
                                     bottom=Side(border_style='thin', color='000000'))
        print('Workbook created, saving to file...')
        wb.save(filename='Diemthi.xlsx')
        os.system("start EXCEL.EXE Diemthi.xlsx")
    except Exception as e:
        print(e)
    print('File saved')
